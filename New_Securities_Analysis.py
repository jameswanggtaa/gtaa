import os
import re
import math
import csv
import time
from concurrent.futures import ThreadPoolExecutor
from typing import Any, Dict, List, Optional, Tuple
from datetime import date

import requests as rq
import pandas as pd
from openpyxl import load_workbook

# ---------------------------------------------------------
# PAC / Proxy‑aware HTTP session (M&T compatible)
# ---------------------------------------------------------
try:
    from pypac import PACSession
except ImportError:
    PACSession = None


def make_http_session():
    """
    Create a requests session that honors corporate proxy/PAC settings.
    If pypac is installed, PACSession will use the system PAC/proxy config.
    Otherwise we fall back to requests.Session (may fail behind proxy).
    """
    if PACSession is not None:
        return PACSession()
    return rq.Session()


def parallel_worker_count() -> int:
    """
    How many securities to process concurrently (thread pool).
    Set YB_WORKERS or YB_PARALLEL_WORKERS (default 1 = sequential).
    Use a modest value (e.g. 4–8); the API may rate-limit heavy parallelism.
    """
    raw = (os.environ.get("YB_WORKERS") or os.environ.get("YB_PARALLEL_WORKERS") or "10").strip()
    try:
        n = int(raw)
        return max(1, min(n, 32))
    except ValueError:
        return 1

# ---------------------------------------------------------
# Configuration
# ---------------------------------------------------------
INPUT_EXCEL = "tba_analysis_input.csv"
INPUT_EXCEL_FALLBACKS = [
    "tba_analysis_input.csv",
    "tba_analysis_input.xlsx",
    "tba_analysis_input_with_cpr.xlsx",
    "tba_analysis_input_old.xlsx",
]
OUTPUT_CSV = "tba_analysis_results.csv"

AUTH_URL = "https://www.yieldbook.com/x/oauth/api-token"
#API_BASE = "https://api.yieldbook.com/analytics/v2"

API_BASE_URL = "https://api.yieldbook.com/analytics/v2"

def api_url(endpoint: str, mode: str | None = None) -> str:
    if not mode:
        return "/".join([API_BASE_URL.strip("/"), endpoint.strip("/")])
    return "/".join([API_BASE_URL.strip("/"), mode.strip("/"), endpoint.strip("/")])


SHOCKS_BPS = [-200, -100, 100, 200]
SCENARIO_TIMING = "Immediate"
SCENARIO_SHIFT_TYPE = "Par"
SCENARIO_INTERPOLATION_TYPE = "Years"
SCENARIO_HORIZON_PY_METHOD = "OAS Change"
SCENARIO_HORIZON_DAYS = 0
SCENARIO_HORIZON_MONTHS = 0


def scenario_horizon_py_method() -> str:
    """
    Horizon repricing method for /bond/scenario-calc (e.g. OAS Change, Spread Change).
    Override with env YB_SCENARIO_HORIZON_PY_METHOD to match Excel/YBSCEN workbook defaults.
    """
    v = (os.environ.get("YB_SCENARIO_HORIZON_PY_METHOD") or "").strip()
    return v if v else SCENARIO_HORIZON_PY_METHOD


def scenario_timing() -> str:
    """Scenario curve shift timing (e.g. Immediate, Gradual). Override: YB_SCENARIO_TIMING."""
    v = (os.environ.get("YB_SCENARIO_TIMING") or "").strip()
    return v if v else SCENARIO_TIMING


def normalize_scenario_horizon(horizon: Any) -> List[Dict[str, Any]]:
    """
    Reorder scenario horizon entries to scen1..scenN matching SHOCKS_BPS when scenarioIDs exist.
    Avoids mis-mapping EffDur columns if the API returns horizons out of order.
    """
    if not isinstance(horizon, list):
        return []
    rows = [h for h in horizon if isinstance(h, dict)]
    if not rows:
        return []
    by_id = {
        str(h["scenarioID"]): h
        for h in rows
        if h.get("scenarioID") is not None and str(h.get("scenarioID")).strip() != ""
    }
    expected = [f"scen{i}" for i in range(1, len(SHOCKS_BPS) + 1)]
    if all(sid in by_id for sid in expected):
        return [by_id[sid] for sid in expected]
    return rows


# Keys for scenario horizon effective duration (Yield Book often populates `duration` only).
SCENARIO_EFFDUR_KEYS = (
    "effectiveDurationAtHorizon",
    "fundedEffectiveDurationAtHorizon",
    "fundedEffectiveDuration",
    "effectiveDuration",
    "durationAtHorizon",
    "duration",
)
MUNI_YBCURVE_TENORS = [
    ("CURRENT", 0.0),
    ("1M", 1.0 / 12.0),
    ("3M", 3.0 / 12.0),
    ("6M", 6.0 / 12.0),
    ("1Y", 1.0),
    ("2Y", 2.0),
    ("3Y", 3.0),
    ("4Y", 4.0),
    ("5Y", 5.0),
    ("7Y", 7.0),
    ("10Y", 10.0),
    ("20Y", 20.0),
    ("30Y", 30.0),
]

OUTPUT_COLUMNS = [
    "CUSIP",
    "Sub Type",

    "Forward_Yield",
    "Prospective_Yield",
    "Effective_Duration",
    "Effective_Convexity",
    "Effective_DV01",
    "Dollar_Duration",

    "PD_1Y",
    "PD_2Y",
    "PD_3Y",
    "PD_5Y",
    "PD_10Y",
    "PD_20Y",
    "PD_30Y",

    "Average_Life",
    "LT_CPR",
    "Life_CPR",
    "OAS",
    "Z_Spread",
    "Factor",
    "GWAC",
    "WALA",
    "WALS",
    "MaxServicerName",
    "MaxServicerPercent",

    "EffDur_-200",
    "DV01_-200",
    "DollarReturn_-200",

    "EffDur_-100",
    "DV01_-100",
    "DollarReturn_-100",

    "EffDur_+100",
    "DV01_+100",
    "DollarReturn_+100",

    "EffDur_+200",
    "DV01_+200",
    "DollarReturn_+200",
]


def scenario_columns() -> List[str]:
    cols: List[str] = []
    for shock in SHOCKS_BPS:
        shock_label = f"{shock:+d}" if shock > 0 else str(shock)
        cols.extend(
            [
                f"EffDur_{shock_label}",
                f"DV01_{shock_label}",
                f"DollarReturn_{shock_label}",
            ]
        )
    return cols
# ---------------------------------------------------------
# Auth
# ---------------------------------------------------------
def load_api_credentials() -> Dict[str, str]:
    """
    Expect credentials from environment variables:
    YB_CLIENT_ID, YB_CLIENT_SECRET
    """
    #return {
    #    "client_id": os.environ["YB_CLIENT_ID"],
    #    "client_secret": os.environ["YB_CLIENT_SECRET"],
    #}
    api_id = "zwang@mtb.com-api"
    api_key = "557ee405-5bc7-f273-5ec4-d9ff91697656"
    return {"client_id": api_id, "client_secret": api_key}



def get_access_token(session) -> str:
    creds = load_api_credentials()
    resp = session.post(
        AUTH_URL,
        data={
            "grant_type": "client_credentials",
            "client_id": creds["client_id"],
            "client_secret": creds["client_secret"],
            "audience": "API2-PROD",
            "ttl": "7200",
        },
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["accessToken"]


def api_headers(token: str) -> Dict[str, str]:
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "*/*",
        "Content-Type": "application/json",
    }

# ---------------------------------------------------------
# Helpers
# ---------------------------------------------------------
def normalize_date(val: Any) -> Optional[str]:
    if val is None or pd.isna(val):
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    return pd.to_datetime(val).strftime("%Y-%m-%d")


def normalize_prepay_type(val: Any) -> Any:
    """
    Map Excel prepay model labels to REST ``prepaySettings.type``.

    Excel's **Prepay Model = Muni** names the municipal *pricing curve* recipe (YBSW),
    not a mortgage PSA/Model stream. For actual REST payloads we normally send
    ``CPY`` / ``0`` instead (see ``prepay_type_and_rate_for_api``); set
    ``YB_MUNI_PREPAY_LEGACY=1`` to restore the older ``Model`` / file-rate mapping.
    """
    if val is None:
        return "Model"
    s = str(val).strip().lower()
    if s in {"", "nan", "none"}:
        return "Model"
    if s == "muni":
        return "Model"
    if s.startswith("model"):
        return int("".join(filter(str.isdigit, s)))
    return s.upper()


def security_uses_municipal_curve(sec: Dict[str, Any]) -> bool:
    """True when Excel prepay model is Muni (YBSW / job-store curve context)."""
    return str(sec.get("prepay_model", "")).strip().lower() == "muni"


def security_is_municipal_bond(sec: Dict[str, Any]) -> bool:
    """
    Municipal product for REST prepay/vol heuristics: prepay model Muni and/or Sub Type MUNI.
    """
    if security_uses_municipal_curve(sec):
        return True
    st = (sec.get("sub_type") or "").strip().upper()
    return "MUNI" in st if st else False


def prepay_type_and_rate_for_api(sec: Dict[str, Any]) -> Tuple[Any, float]:
    """
    ``prepaySettings`` type and rate for /bond/py, scenario-calc, and keyword py.

    Municipal bonds: default **CPY / 0** (same convention as Agency passthrough rows in
    many portfolio exports). Excel often shows Prepay Model Muni with Prepay Rate 100;
    sending **Model / 100** to REST can look like a fully ramped mortgage model and
    leave option-sensitive fields (OAS, WAL, CPR paths) empty.

    Override defaults:
    - ``YB_MUNI_PREPAY_LEGACY=1`` — use ``normalize_prepay_type`` + file prepay rate.
    - ``YB_MUNI_CPY_RATE`` — if set (numeric), use **CPY** with that rate instead of 0.
    """
    if sub_type_is_treasury(sec.get("sub_type")):
        pr = sec.get("prepay_rate")
        if pr is None:
            pr = 0.0
        try:
            prf = float(pr)
        except (TypeError, ValueError):
            pn = parse_number(pr)
            prf = float(pn) if pn is not None else 0.0
        return normalize_prepay_type(sec.get("prepay_model")), prf

    legacy = (os.environ.get("YB_MUNI_PREPAY_LEGACY") or "").strip().lower() in {"1", "true", "yes", "y"}
    if not legacy and security_is_municipal_bond(sec):
        cpy_custom = parse_number(os.environ.get("YB_MUNI_CPY_RATE"))
        rate = float(cpy_custom) if cpy_custom is not None else 0.0
        return "CPY", rate

    pr = sec.get("prepay_rate")
    if pr is None:
        pr = 100.0
    try:
        prf = float(pr)
    except (TypeError, ValueError):
        pn = parse_number(pr)
        prf = float(pn) if pn is not None else 100.0
    return normalize_prepay_type(sec.get("prepay_model")), prf


def py_calc_props(sec: Dict[str, Any]) -> Dict[str, str]:
    """REST ``props`` for pyCalcInputs / sync input (subtype is not a top-level field)."""
    st = (sec.get("sub_type") or "").strip()
    return {"subType": st} if st else {}


def yieldbook_security_identifier(sec: Dict[str, Any]) -> str:
    """
    REST ``identifier`` for ``securityIDEntry`` (/bond/py, scenario-calc, indic, async py, …).

    Many **Structured Agency CMBS** tranches are stored in Yield Book under ``<CUSIP>.cmo``.
    Workbooks often list the nine-character CUSIP only; with that form, sync PY can return
    empty yields and durations while still returning a few collateral fields. Appending
    ``.cmo`` (when not already present) restores full pricing for those names. Other sub types
    are returned unchanged (``Agency CMBS`` / ``Agency CMO`` rows behave with or without the suffix
    in typical tenants).
    """
    c = str(sec.get("cusip") or "").strip()
    if not c or c.upper().endswith(".CMO"):
        return c
    st = (sec.get("sub_type") or "").strip().lower()
    if st == "structured agency cmbs":
        return f"{c}.cmo"
    return c


def muni_rest_embed_curve_points_enabled() -> bool:
    """
    When False, municipal REST requests omit embedded tenor/rate pillars (see ``curve_dict_for``).

    For **Prepay Model Muni** without a job-store ``$ref``, the curve is then **``{curveType}`` only**
    (e.g. ``SWAP_RFR``), matching Excel ``YBPRICE(..., \"RFRSwap\", ...)`` / LSEG samples.

    Set ``YB_MUNI_INCLUDE_CURVE_POINTS=0`` to disable embedding. Default is on (1) for job-store / tenants that accept inline spots.
    """
    raw = (os.environ.get("YB_MUNI_INCLUDE_CURVE_POINTS") or "1").strip().lower()
    return raw in {"1", "true", "yes", "y"}


def curve_type(val: Any, prepay_model: Any) -> str:
    """
    REST curveType string for Yield Book (maps file labels like RFRSwap to API enums).

    For prepay model Muni, default to SWAP_RFR because many tenants reject
    municipal enum labels in REST curveType validation.
    Override with YB_MUNI_CURVE_TYPE if your tenant supports a municipal enum.
    """
    if str(prepay_model).strip().lower() == "muni":
        ct = (os.environ.get("YB_MUNI_CURVE_TYPE") or "SWAP_RFR").strip()
        return ct if ct else "SWAP_RFR"
    s = str(val).strip().upper()
    if "RFR" in s:
        return "SWAP_RFR"
    return s or "SWAP_RFR"


def curve_dict_for(sec: Dict[str, Any]) -> Dict[str, Any]:
    """
    Full curve object for /bond/py and scenario-calc.

    Excel YBSW(\"MUNI\",\"USD\", tenors, rates, ...) is evaluated in the add-in only; REST
    does not expose a YBSW() RPC. To align with that curve you must either use a
    documented curveType from LSEG or pass user-defined spots if/when your contract's
    OpenAPI lists the exact field names (trial \"MUNI\" as curveType returned HTTP 400).

    Optional tenor/rate pairs are stored on the security as muni_curve_points (from
    Excel B2:B14/E2:E14 or muni_ybsw_curve.csv) for logging and for future custom-curve
    wiring once the schema is known.

    When LSEG support's job-store user curve is used, ``sec[\"yb_user_curve_ref\"]`` holds
    the upload request id (e.g. R-733194743). See ``prepare_muni_job_store_curve_ref`` and
    set env ``YB_MUNI_JOB_STORE=1`` before running muni benchmarks.

    Embedded pillars are omitted when ``muni_rest_embed_curve_points_enabled()`` is false
    (``YB_MUNI_INCLUDE_CURVE_POINTS=0``): for **Prepay Model Muni** (no job-store ``$ref``)
    the curve is **``{curveType}`` only** (e.g. ``SWAP_RFR``), matching Excel
    ``YBPRICE(..., \"RFRSwap\", ...)`` / LSEG sync samples — no ``currency``, no muni pillars.

    With embedded pillars or ``yb_user_curve_ref``, shape follows job-store / tenant rules.
    """
    ref = str(sec.get("yb_user_curve_ref") or "").strip()
    if ref and security_uses_municipal_curve(sec):
        return {"userDefined": {"$ref": ref}}

    ct = curve_type(sec.get("curve_type"), sec.get("prepay_model"))
    is_muni = security_uses_municipal_curve(sec)
    use_custom = is_muni and muni_rest_embed_curve_points_enabled()

    # Optional: attach YBSW-style muni tenor/rate points into curve object.
    # This is tenant/schema dependent; keep behind an env flag.
    if use_custom:
        pts_raw = sec.get("muni_curve_points") or []
        pts: List[Dict[str, float]] = []
        for p in pts_raw:
            if not isinstance(p, dict):
                continue
            y = parse_number(p.get("year"))
            r = parse_number(p.get("rate"))
            if y is None or r is None:
                continue
            pts.append({"year": float(y), "rate": float(r)})

        if pts:
            curve_obj: Dict[str, Any] = {"curveType": ct, "currency": "USD"}
            # Field names are configurable because contracts may use different schemas.
            # Examples to try:
            #   YB_MUNI_CURVE_POINTS_KEYS=curveSpots,userCurveSpots
            #   YB_MUNI_CURVE_YEAR_KEY=year
            #   YB_MUNI_CURVE_RATE_KEY=value
            keys_raw = (os.environ.get("YB_MUNI_CURVE_POINTS_KEYS") or "curveSpots").strip()
            keys = [k.strip() for k in keys_raw.split(",") if k.strip()]
            year_key = (os.environ.get("YB_MUNI_CURVE_YEAR_KEY") or "year").strip()
            rate_key = (os.environ.get("YB_MUNI_CURVE_RATE_KEY") or "rate").strip()
            mapped_pts = [{year_key: p["year"], rate_key: p["rate"]} for p in pts]
            for k in keys:
                curve_obj[k] = mapped_pts
            return curve_obj

    if is_muni:
        return {"curveType": ct}

    return {"curveType": ct, "currency": "USD"}


def extract_yb_job_store_ref(obj: Any) -> Optional[str]:
    """Locate R-######## request id from analytics v2 job-store JSON (shape varies by tenant)."""
    if obj is None:
        return None
    if isinstance(obj, str):
        s = obj.strip()
        if re.fullmatch(r"R-\d+", s):
            return s
        return None
    if isinstance(obj, dict):
        for k in ("requestId", "request_id", "id", "resourceId", "curveRequestId", "requestID"):
            v = obj.get(k)
            if isinstance(v, str):
                vs = v.strip()
                if re.fullmatch(r"R-\d+", vs):
                    return vs
        for v in obj.values():
            found = extract_yb_job_store_ref(v)
            if found:
                return found
    if isinstance(obj, list):
        for v in obj:
            found = extract_yb_job_store_ref(v)
            if found:
                return found
    return None


def create_analytics_job(session, token: str, job_name: str) -> Dict[str, Any]:
    """POST /analytics/v2/jobs/ — container for user curve uploads (LSEG muni workflow)."""
    url = api_url("jobs").rstrip("/") + "/"
    r = session.post(url, json={"name": job_name}, headers=api_headers(token), timeout=60)
    if not r.ok:
        snippet = (r.text or "")[:1200]
        raise RuntimeError(f"create job HTTP {r.status_code}: {snippet}")
    if not (r.text or "").strip():
        return {}
    try:
        return r.json()
    except Exception:
        return {}


def upload_job_store_user_curve(
    session,
    token: str,
    job_name: str,
    term_rate_points: List[Dict[str, float]],
    asof: str,
) -> str:
    """
    POST /analytics/v2/job-store/{job}/curve with USERSW-style swap curve (support recipe for muni pillars).

    term_rate_points: items with float \"term\" (year fraction) and \"rate\" (percent, same as Excel YBSW column).
    """
    if not term_rate_points:
        raise ValueError("term_rate_points is empty")
    swap_index = (os.environ.get("YB_MUNI_JOB_SWAP_INDEX") or "SOFR").strip() or "SOFR"
    try:
        compound = int(os.environ.get("YB_MUNI_JOB_COMPOUNDING_FREQ") or "2")
    except ValueError:
        compound = 2
    body: Dict[str, Any] = {
        "curveType": "SWAP",
        "currency": "USD",
        "points": [{"rate": float(p["rate"]), "term": float(p["term"])} for p in term_rate_points],
        "rateType": "Par",
        "source": "USERSW",
        "swapIndex": swap_index,
        "asof": asof,
        "compoundingFreq": compound,
    }
    url = api_url(f"job-store/{job_name}/curve")
    r = session.post(url, json=body, headers=api_headers(token), timeout=120)
    if not r.ok:
        snippet = (r.text or "")[:2000]
        raise RuntimeError(f"job-store curve HTTP {r.status_code}: {snippet}")
    ref = extract_yb_job_store_ref(r.json())
    if not ref:
        snippet = (r.text or "")[:2000]
        raise ValueError(f"job-store curve upload succeeded but no R- request id in JSON: {snippet}")
    return ref


def muni_points_to_job_store_terms(points: List[Dict[str, float]]) -> List[Dict[str, float]]:
    """Map internal {\"year\", \"rate\"} muni pillars to job-store {\"term\", \"rate\"}."""
    out: List[Dict[str, float]] = []
    for p in points:
        if not isinstance(p, dict):
            continue
        y = parse_number(p.get("year"))
        r = parse_number(p.get("rate"))
        if y is None or r is None:
            continue
        term = float(y)
        if term <= 0.0:
            term = float(os.environ.get("YB_MUNI_JOB_ZERO_TERM") or "0.002739726")
        out.append({"term": term, "rate": float(r)})
    return out


def prepare_muni_job_store_curve_ref(session, token: str, securities: List[Dict[str, Any]]) -> None:
    """
    One shared user curve for all Muni prepay_model rows (matches LSEG job-store guidance).

    Enable with: YB_MUNI_JOB_STORE=1
    Optional: YB_MUNI_JOB_NAME=mycurve  (default: muni + millisecond timestamp)
    """
    flag = (os.environ.get("YB_MUNI_JOB_STORE") or "").strip().lower() in {"1", "true", "yes", "y"}
    if not flag:
        return
    muni_secs = [s for s in securities if security_uses_municipal_curve(s)]
    if not muni_secs:
        return
    first = muni_secs[0]
    first["muni_curve_points"] = resolve_muni_curve_points(session, token, first)
    pts = first.get("muni_curve_points") or []
    if not pts:
        print(
            "[WARN] YB_MUNI_JOB_STORE: no muni curve points; skipping job-store upload. "
            "Add muni_ybsw_curve.csv or muni_ybcurve.csv beside the portfolio file, "
            "or put tenor/rate pillars in the Muni .xlsx (B2:B14 / E2:E14)."
        )
        return
    for s in muni_secs:
        s["muni_curve_points"] = list(pts)
    asof = str(first.get("curve_date") or default_pricing_date())
    job_name = (os.environ.get("YB_MUNI_JOB_NAME") or "").strip() or f"muni{int(time.time() * 1000)}"
    term_rates = muni_points_to_job_store_terms(pts)
    if not term_rates:
        print("[WARN] YB_MUNI_JOB_STORE: could not map muni_curve_points to term/rate; skipping.")
        return
    try:
        create_analytics_job(session, token, job_name)
    except Exception as e:
        print(f"[WARN] YB_MUNI_JOB_STORE: create job ({job_name!r}) failed: {e}; retrying alternate name.")
        job_name = f"{job_name}_{int(time.time() * 1000)}"
        create_analytics_job(session, token, job_name)
    ref = upload_job_store_user_curve(session, token, job_name, term_rates, asof)
    for s in muni_secs:
        s["yb_user_curve_ref"] = ref
    print(
        f"[INFO] YB_MUNI_JOB_STORE: user curve ref={ref} job={job_name!r} "
        f"points={len(term_rates)} asof={asof}"
    )


def _extract_curve_points_from_obj(obj: Any) -> List[Dict[str, float]]:
    """Best-effort parser for curve point arrays in unknown tenant response shapes."""
    out: List[Dict[str, float]] = []
    if not isinstance(obj, list):
        return out
    for item in obj:
        if not isinstance(item, dict):
            continue
        y = parse_number(get_first_key(item, "year", "tenorYear", "tenor", "maturity", "x"))
        r = parse_number(get_first_key(item, "rate", "value", "spot", "parRate", "y"))
        if y is None or r is None:
            continue
        out.append({"year": float(y), "rate": float(r)})
    return out


def fetch_muni_curve_points_from_ybcurve(session, token: str, sec: Dict[str, Any]) -> List[Dict[str, float]]:
    """
    Try to pull YBCURVE tenor/rate points from keyword PY request.
    If unavailable for the tenant/schema, return [].
    """
    pricing_date = sec.get("curve_date") or default_pricing_date()
    try:
        lvl = str(bond_py_level_value(sec))
    except ValueError:
        return []
    body = {
        "keywords": ["YBCURVE"],
        "globalSettings": {"pricingDate": pricing_date},
        "pyCalcInputs": [{
            "identifier": sec["cusip"],
            "idType": "securityIDEntry",
            "level": lvl,
            "settlementDate": pricing_date,
            "curve": {"curveType": "SWAP_RFR", "currency": "USD"},
            "prepaySettings": {"type": "Model", "rate": sec.get("prepay_rate") or 100.0},
            "volatility": {"type": "MatrixWSkew"},
        }],
    }
    r = session.post(api_url("bond/py", mode="req"), json=body, headers=api_headers(token), timeout=60)
    if not r.ok:
        return []
    request_id = r.json().get("requestId")
    if not request_id:
        return []

    results_url = api_url(f"/results/{request_id}", mode=None)
    for _ in range(20):
        rr = session.get(results_url, headers=api_headers(token), timeout=30)
        if rr.status_code == 404:
            time.sleep(1)
            continue
        if not rr.ok:
            return []
        jr = rr.json()
        if jr.get("meta", {}).get("status") == "DONE":
            py_kw = (jr.get("results") or [{}])[0].get("py", {}) or {}
            if not isinstance(py_kw, dict):
                return []
            for k, v in py_kw.items():
                ku = str(k).lower()
                if "ybcurve" in ku or ("curve" in ku and isinstance(v, list)):
                    pts = _extract_curve_points_from_obj(v)
                    if pts:
                        return pts
            return []
        time.sleep(1)
    return []


def resolve_muni_curve_points(session, token: str, sec: Dict[str, Any]) -> List[Dict[str, float]]:
    """
    Resolve Muni curve points in priority order:
    1) Existing concrete points on sec
    2) YBCURVE keyword pull (tenant-dependent)
    3) []
    """
    pts_raw = sec.get("muni_curve_points") or []
    concrete: List[Dict[str, float]] = []
    for p in pts_raw:
        if not isinstance(p, dict):
            continue
        y = parse_number(p.get("year"))
        r = parse_number(p.get("rate"))
        if y is None or r is None:
            continue
        concrete.append({"year": float(y), "rate": float(r)})
    if concrete:
        return concrete
    return fetch_muni_curve_points_from_ybcurve(session, token, sec)


def sub_type_uses_bond_yield(sub_type: Optional[str]) -> bool:
    """Muni and Treasury rows use street/yield; others use forward yield from the API."""
    st = (sub_type or "").strip().upper()
    if not st:
        return False
    if "MUNI" in st:
        return True
    if "TREASUR" in st:
        return True
    return False


def sub_type_is_treasury(sub_type: Optional[str]) -> bool:
    """Treasuries have no prepayment / vol model in the add-in; input fields stay blank."""
    st = (sub_type or "").strip().upper()
    return bool(st and "TREASUR" in st)


def resolve_forward_yield_column(py: Dict[str, Any], sub_type: Optional[str]) -> Any:
    """
    OUTPUT column Forward_Yield mapping rule:
    - Muni/Treasury: use bond yield (py.yield / effectiveYield)
    - All others (Agency MBS/CMO/etc.): use forward yield (forwardMeasures.yield)
    """
    fy = py.get("forwardYield")
    if fy is None:
        fwd_measures = py.get("forwardMeasures") or {}
        if isinstance(fwd_measures, dict):
            fy = get_first_key(fwd_measures, "yield", "Yield")
    y = get_first_key(py, "bondYield", "yield", "effectiveYield", "streetYield", "Yield")
    if sub_type_uses_bond_yield(sub_type):
        return y
    # For non-Muni/Treasury, prefer forward yield; fallback to other yield fields if missing.
    return fy if fy is not None else y


def resolve_prospective_yield_from_py(py_at_book: Optional[Dict[str, Any]], sub_type: Optional[str]) -> Any:
    """
    OUTPUT column Prospective_Yield: same measure as ``Forward_Yield`` (Excel YBPRICE ``forwardYield`` slot),
    but from a **book-price** /bond/py pass (``MTB_BOOK PRICE (Clean)``), matching workbook Prospective Yield.
    """
    if not py_at_book:
        return pd.NA
    return resolve_forward_yield_column(py_at_book, sub_type)


def muni_vol_single_from_curve_enabled() -> bool:
    """
    When true, municipal (Prepay Model Muni) rows use ``volatility.type`` = ``Single`` with
    ``volatility.rate`` interpolated from ``muni_curve_points`` (sidecar CSV / job-store pillars),
    while ``YB_MUNI_INCLUDE_CURVE_POINTS=0`` can still keep the **curve** object as plain ``SWAP_RFR``.

    Set ``YB_MUNI_VOL_SINGLE_FROM_CURVE=1``. Optional: ``YB_MUNI_VOL_INTERP_YEAR`` (default **10**)
    for the pillar year used in linear interpolation.

    This is an experimental knob for benchmark research vs ``MatrixWSkew`` defaults.
    """
    return (os.environ.get("YB_MUNI_VOL_SINGLE_FROM_CURVE") or "").strip().lower() in {"1", "true", "yes", "y"}


def derive_volatility_rate_from_muni_points(
    pts: List[Dict[str, Any]],
    target_year: Optional[float] = None,
) -> Optional[float]:
    """
    Linearly interpolate **par rate** (percent units as stored in sidecar) at ``target_year``
    from ``muni_curve_points`` ``{year, rate}`` nodes. Used only as a scalar ``Single`` vol input
    when exploring REST behavior; it is **not** guaranteed to match Yield Book OAS vol conventions.
    """
    if target_year is None:
        ty_raw = (os.environ.get("YB_MUNI_VOL_INTERP_YEAR") or "10").strip()
        try:
            target_year = float(ty_raw)
        except ValueError:
            target_year = 10.0

    pairs: List[Tuple[float, float]] = []
    for p in pts:
        if not isinstance(p, dict):
            continue
        y = parse_number(p.get("year"))
        r = parse_number(p.get("rate"))
        if y is None or r is None:
            continue
        pairs.append((float(y), float(r)))
    if not pairs:
        return None
    pairs.sort(key=lambda t: t[0])
    if len(pairs) == 1:
        return pairs[0][1]
    y_t = float(target_year)
    for y, r in pairs:
        if abs(y - y_t) < 1e-9:
            return float(r)
    if y_t <= pairs[0][0]:
        return float(pairs[0][1])
    if y_t >= pairs[-1][0]:
        return float(pairs[-1][1])
    for i in range(len(pairs) - 1):
        y0, r0 = pairs[i]
        y1, r1 = pairs[i + 1]
        if y0 <= y_t <= y1:
            if abs(y1 - y0) < 1e-15:
                return float(r0)
            t = (y_t - y0) / (y1 - y0)
            return float(r0 + t * (r1 - r0))
    return None


def resolve_volatility(sec: Dict[str, Any]) -> Dict[str, Any]:
    """
    Build volatility payload aligned to product conventions.
    For Muni, Excel-style runs commonly use MatrixWSkew when input shows Single.
    """
    if sub_type_is_treasury(sec.get("sub_type")):
        return {"type": "Default"}

    vol_raw = sec.get("vol_model")
    vol_model = (str(vol_raw).strip() if vol_raw is not None else "") or "Default"
    is_muni = security_is_municipal_bond(sec)
    vol_upper = vol_model.upper()

    if muni_vol_single_from_curve_enabled() and security_uses_municipal_curve(sec):
        pts = sec.get("muni_curve_points") or []
        vr = derive_volatility_rate_from_muni_points(pts)
        if vr is not None:
            # API accepts string or number; keep short decimal string.
            rate_s = f"{float(vr):.6f}".rstrip("0").rstrip(".")
            return {"type": "Single", "rate": rate_s if rate_s else "0"}

    if is_muni and vol_upper == "SINGLE":
        return {"type": "MatrixWSkew"}
    if vol_upper == "SINGLE":
        return {"type": "Single", "rate": 0}
    return {"type": vol_model}


def computed_dollar_duration(sec: Dict[str, Any], py: Dict[str, Any]) -> Any:
    """
    Dollar_Duration = current_factor * Nominal * effectiveDV01 / 100
    current_factor defaults to 1 when not in input; Nominal from input column.
    Uses formula-only output (no API dollarDuration fallback).
    """
    dv01 = py.get("effectiveDV01")
    if dv01 is None:
        return None
    try:
        dv01_f = float(dv01)
    except (TypeError, ValueError):
        return None

    cf_raw = sec.get("current_factor")
    if cf_raw is None:
        cf = 1.0
    else:
        try:
            cf = float(cf_raw)
        except (TypeError, ValueError):
            cf = 1.0

    nom = sec.get("nominal")
    if nom is None:
        return None
    try:
        nom_f = float(nom)
    except (TypeError, ValueError):
        return None

    return cf * nom_f * dv01_f / 100.0


def pick_input_file() -> str:
    if os.path.isfile(INPUT_EXCEL):
        return INPUT_EXCEL
    for path in INPUT_EXCEL_FALLBACKS:
        if os.path.isfile(path):
            return path
    raise FileNotFoundError(
        "No input file found. Expected one of: "
        + ", ".join(INPUT_EXCEL_FALLBACKS)
    )


def normalize_colname(s: Any) -> str:
    return str(s).strip().lower().replace(" ", "").replace("_", "")


def col_lookup(df: pd.DataFrame) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for c in df.columns:
        out[normalize_colname(c)] = c
    return out


def find_col(cols: Dict[str, str], *names: str) -> Optional[str]:
    for n in names:
        key = normalize_colname(n)
        if key in cols:
            return cols[key]
    # Handle frequent typo: curve_trpe
    if any(normalize_colname(n) in {"curvetrpe", "curvetype"} for n in names):
        for key, actual in cols.items():
            if "curve" in key and ("type" in key or "trpe" in key):
                return actual
    return None


def parse_number(val: Any) -> Optional[float]:
    if val is None or pd.isna(val):
        return None
    s = str(val).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def bond_py_level_value(sec: Dict[str, Any]) -> float:
    """
    Numeric market level for Yield Book bond/py (field ``level``).
    Raises ValueError if missing or non-finite; the API rejects e.g. ``Invalid level: nan``.
    """
    mp = sec.get("market_price")
    if mp is None or pd.isna(mp):
        raise ValueError(
            "Missing or invalid market price (Yield Book requires a numeric level / price)."
        )
    try:
        v = float(mp)
    except (TypeError, ValueError):
        pn = parse_number(mp)
        if pn is None or pd.isna(pn):
            raise ValueError(
                "Missing or invalid market price (Yield Book requires a numeric level / price)."
            )
        v = float(pn)
    if math.isnan(v) or math.isinf(v):
        raise ValueError(
            "Missing or invalid market price (Yield Book requires a numeric level / price)."
        )
    return v


def bond_py_level_for_request(sec: Dict[str, Any], level_override: Optional[float] = None) -> float:
    """
    Numeric ``level`` for /bond/py. Defaults to ``market_price`` (``bond_py_level_value``).
    ``level_override`` is used for a second pass at **book price** (Excel Prospective Yield / YBPRICE book).
    """
    if level_override is not None:
        if isinstance(level_override, float) and pd.isna(level_override):
            raise ValueError("Invalid level_override (nan) for Yield Book /bond/py level.")
        try:
            v = float(level_override)
        except (TypeError, ValueError):
            pn = parse_number(level_override)
            if pn is None or (isinstance(pn, float) and pd.isna(pn)):
                raise ValueError("Invalid level_override for Yield Book /bond/py level.")
            v = float(pn)
        if math.isnan(v) or math.isinf(v):
            raise ValueError("Invalid level_override for Yield Book /bond/py level.")
        return v
    return bond_py_level_value(sec)


def clean_text(val: Any) -> Optional[str]:
    if val is None or pd.isna(val):
        return None
    s = str(val).strip()
    return s or None


def get_first_key(d: Dict[str, Any], *names: str) -> Any:
    """
    Return the first non-None value from candidate keys, case-insensitive.
    """
    if not isinstance(d, dict) or not d:
        return None
    lower_map = {str(k).lower(): v for k, v in d.items()}
    for name in names:
        if name in d and d.get(name) is not None:
            return d.get(name)
        v = lower_map.get(str(name).lower())
        if v is not None:
            return v
    return None


def extract_pphist_cpr_life(payload: Dict[str, Any]) -> Any:
    """
    Extract Life CPR from either flat keys or nested PPM history list shape:
    dataPPMHistoryList -> prepayType=CPR -> dataPPMHistoryDetailList -> month=Life.
    """
    direct = get_first_key(payload, "ppHistCPRLife", "PPHistCPRLife", "lifeCPR", "LifeCPR")
    if direct is not None:
        return direct

    history_list = get_first_key(payload, "dataPPMHistoryList", "datappmhistorylist")
    if not isinstance(history_list, list):
        return None

    for block in history_list:
        if not isinstance(block, dict):
            continue
        prepay_type = str(get_first_key(block, "prepayType", "prepaytype") or "").strip().upper()
        if prepay_type != "CPR":
            continue
        details = get_first_key(block, "dataPPMHistoryDetailList", "datappmhistorydetaillist")
        if not isinstance(details, list):
            continue
        for item in details:
            if not isinstance(item, dict):
                continue
            month = str(get_first_key(item, "month", "Month") or "").strip().upper()
            if month == "LIFE":
                return get_first_key(item, "prepayRate", "prepayrate", "rate", "value")
    return None


def extract_long_term_cpr(payload: Dict[str, Any]) -> Any:
    """
    Extract LongTerm CPR from either flat keys or nested projection list shape:
    dataPpmProjList -> prepayType=CPR -> longTerm.
    """
    direct = get_first_key(payload, "longTermCPR", "LongTermCPR", "ltCPR", "LTCPR")
    if direct is not None:
        return direct

    proj_list = get_first_key(payload, "dataPpmProjList", "dataPPMProjList", "datappmprojlist")
    if not isinstance(proj_list, list):
        return None

    for block in proj_list:
        if not isinstance(block, dict):
            continue
        prepay_type = str(get_first_key(block, "prepayType", "prepaytype") or "").strip().upper()
        if prepay_type != "CPR":
            continue
        return get_first_key(block, "longTerm", "longterm", "lt", "value")
    return None


def resolve_life_cpr_value(sub_type: Optional[str], py: Dict[str, Any], indic: Dict[str, Any]) -> Any:
    """
    Life CPR selection:
    1) PY Life CPR
    2) INDIC Life CPR
    3) For CMO-style sectors, fallback to CPR long-term projection as proxy
    """
    py_life = py.get("PPHistCPRLife")
    if py_life is not None:
        return py_life
    indic_life = indic.get("PPHistCPRLife")
    if indic_life is not None:
        return indic_life

    st = (sub_type or "").strip().upper()
    if "CMO" in st:
        py_lt = py.get("LongTermCPR")
        if py_lt is not None:
            return py_lt
        indic_lt = indic.get("LongTermCPR")
        if indic_lt is not None:
            return indic_lt
    return None


def default_pricing_date() -> str:
    return date.today().strftime("%Y-%m-%d")


def read_muni_curve_points(input_file: str) -> List[Dict[str, float]]:
    """
    Read tenor/rate columns that feed Excel YBSW(\"MUNI\",\"USD\", $B$2:$B$14, $E$2:$E$14, ...).

    From .xlsx: active sheet B2:B14 (years) and E2:E14 (rates, as in the add-in).
    If the input is not xlsx or ranges are empty, tries read_muni_spot_curve_sidecar().
    """
    if str(input_file).lower().endswith(".xlsx"):
        try:
            wb = load_workbook(input_file, data_only=True, read_only=True)
            ws = wb.active
            points: List[Dict[str, float]] = []
            for r in range(2, 15):
                tenor = parse_number(ws[f"B{r}"].value)
                rate = parse_number(ws[f"E{r}"].value)
                if tenor is None or rate is None:
                    continue
                points.append({"year": float(tenor), "rate": float(rate)})
            wb.close()
            if points:
                return points
        except Exception:
            pass
    return read_muni_spot_curve_sidecar(input_file)


def default_muni_ybcurve_points() -> List[Dict[str, Any]]:
    """
    YBCURVE tenor skeleton:
    current, 1m, 3m, 6m, 1y, 2y, 3y, 4y, 5y, 7y, 10y, 20y, 30y.
    """
    return [{"tenor": t, "year": y, "rate": None} for t, y in MUNI_YBCURVE_TENORS]


def read_muni_spot_curve_sidecar(input_file: str) -> List[Dict[str, float]]:
    """
    Optional CSV alongside the portfolio file (same directory as input_file):

      muni_ybcurve.csv     (or Muni_ybcurve.csv)
      muni_ybsw_curve.csv  (or Muni_ybsw_curve.csv)

    Columns (case-insensitive): YrLookup (Excel YBSW column B) or year/tenor/term, plus
    rate (or parrate, spot, yield). Optional: CurveDate, Years (labels like 1 Month), Curve.
    Accepts YBCURVE-style tenor labels:
      current, 1m, 3m, 6m, 1y, 2y, 3y, 4y, 5y, 7y, 10y, 20y, 30y
    """
    tenor_years = {k.lower(): v for k, v in MUNI_YBCURVE_TENORS}

    def _parse_tenor_to_years(raw: Any) -> Optional[float]:
        if raw is None or pd.isna(raw):
            return None
        s = str(raw).strip().lower().replace(" ", "")
        if not s:
            return None
        if s in tenor_years:
            return float(tenor_years[s])
        if s.endswith("month"):
            s = s.replace("month", "m")
        if s.endswith("months"):
            s = s.replace("months", "m")
        if s.endswith("year"):
            s = s.replace("year", "y")
        if s.endswith("years"):
            s = s.replace("years", "y")
        if s in tenor_years:
            return float(tenor_years[s])
        parsed_num = parse_number(s)
        return float(parsed_num) if parsed_num is not None else None

    d = os.path.dirname(os.path.abspath(input_file))
    for name in ("muni_ybcurve.csv", "Muni_ybcurve.csv", "muni_ybsw_curve.csv", "Muni_ybsw_curve.csv"):
        p = os.path.join(d, name)
        if not os.path.isfile(p):
            continue
        try:
            df = pd.read_csv(p)
            cols = col_lookup(df)
            ycol = find_col(
                cols,
                "yrlookup",
                "year",
                "tenor",
                "maturity",
                "maturityyears",
                "term",
                "years",
            )
            rcol = find_col(cols, "parrate", "rate", "spot", "yield", "par", "zero")
            if not ycol or not rcol:
                continue
            out: List[Dict[str, float]] = []
            for _, row in df.iterrows():
                yv = _parse_tenor_to_years(row[ycol])
                rv = parse_number(row[rcol])
                if yv is None or rv is None:
                    continue
                out.append({"year": float(yv), "rate": float(rv)})
            return out
        except Exception:
            continue
    return []


def muni_curve_points_debug_json(sec: Dict[str, Any]) -> str:
    """Compact JSON of muni_curve_points for support / OpenAPI questions."""
    import json

    pts = sec.get("muni_curve_points") or []
    return json.dumps(pts, indent=2)

# ---------------------------------------------------------
# Load Excel
# ---------------------------------------------------------
def load_securities() -> List[Dict[str, Any]]:
    input_file = pick_input_file()
    if input_file.lower().endswith(".csv"):
        df = pd.read_csv(input_file)
    else:
        df = pd.read_excel(input_file)

    cols = col_lookup(df)
    cusip_col = find_col(cols, "CUSIP", "CUSIPs")
    coupon_col = find_col(cols, "Coupon")
    maturity_col = find_col(cols, "Maturity_Date", "Maturity Date", "Maturity")
    sub_type_col = find_col(cols, "Sub Type", "Sub_Type", "Subtype")
    market_price_col = find_col(cols, "Market_price", "Market Price")
    book_price_col = find_col(
        cols,
        "MTB_BOOK PRICE (Clean)",
        "MTB_BOOK_PRICE",
        "MTB Book Price (Clean)",
        "Book Price",
    )
    curve_date_col = find_col(cols, "Curve_Date", "Curve date")
    prepay_model_col = find_col(cols, "Prepay_Model", "Prepayment model", "Prepayment_Model")
    prepay_rate_col = find_col(cols, "Prepay_Rate", "prepay_rate")
    vol_model_col = find_col(cols, "Vol_Model", "vol_model")
    curve_type_col = find_col(cols, "Curve_Type", "curve_type", "curve_trpe")
    nominal_col = find_col(cols, "Nominal", "CA_NOTIONAL", "Notional", "Current Balance")
    current_factor_col = find_col(cols, "Current Factor", "Current_Factor", "current_factor", "Factor_Input")

    required = {
        "CUSIP": cusip_col,
        "Coupon": coupon_col,
        "Maturity_Date": maturity_col,
        "Market_price": market_price_col,
    }
    missing = [k for k, v in required.items() if v is None]
    if missing:
        raise ValueError(f"Missing required column(s): {missing}. Found columns: {list(df.columns)}")

    muni_curve_points = read_muni_curve_points(input_file)
    if not muni_curve_points:
        muni_curve_points = default_muni_ybcurve_points()
    out = []

    for _, r in df.iterrows():
        sub_type = clean_text(r[sub_type_col]) if sub_type_col else None
        prepay_model = r[prepay_model_col] if prepay_model_col else None
        cf_val = parse_number(r[current_factor_col]) if current_factor_col else None
        if sub_type_is_treasury(sub_type):
            prepay_model = None
            prepay_rate_val = None
            vol_model_val = None
        else:
            prepay_rate_val = parse_number(r[prepay_rate_col]) if prepay_rate_col else 100.0
            vol_model_val = clean_text(r[vol_model_col]) if vol_model_col else None
        out.append({
            "cusip": str(r[cusip_col]).strip(),
            "sub_type": sub_type,
            "coupon": parse_number(r[coupon_col]) if coupon_col else None,
            "maturity": normalize_date(r[maturity_col]),
            "market_price": parse_number(r[market_price_col]),
            "book_price": parse_number(r[book_price_col]) if book_price_col else None,
            "curve_date": normalize_date(r[curve_date_col]) if curve_date_col else default_pricing_date(),
            "prepay_model": prepay_model,
            "prepay_rate": prepay_rate_val,
            "vol_model": vol_model_val,
            "curve_type": clean_text(r[curve_type_col]) if curve_type_col else "SWAP_RFR",
            "nominal": parse_number(r[nominal_col]) if nominal_col else None,
            "current_factor": float(cf_val) if cf_val is not None else 1.0,
            "muni_curve_points": muni_curve_points if str(prepay_model).strip().lower() == "muni" else [],
        })
    return out

# ---------------------------------------------------------
# PY Analytics
# ---------------------------------------------------------
def muni_sync_py_ybprice_rfrswap_shape(sec: Dict[str, Any], curve_obj: Dict[str, Any]) -> bool:
    """
    True when sync ``/bond/py`` should match Excel ``YBPRICE(..., \"RFRSwap\", ...)``:
    only ``curveType`` on the curve (no ``currency``, pillars, or job-store ``$ref``).

    If ``curve_obj`` includes embedded spot arrays or ``userDefined``, the full pyCalc
    shape (``prepaySettings``, ``idType``, …) is used instead.
    """
    if not security_uses_municipal_curve(sec):
        return False
    if str(sec.get("yb_user_curve_ref") or "").strip():
        return False
    return set(curve_obj.keys()) == {"curveType"}


def structured_agency_cmbs_ybprice_sync_shape(sec: Dict[str, Any], curve_obj: Dict[str, Any]) -> bool:
    """
    True when sync ``/bond/py`` should match the Excel add-in for **Structured Agency CMBS**:
    ``CUSIP.cmo`` id, ``curveType`` = SWAP_RFR only, ``prepaySettings`` (e.g. CPY / 0),
    ``volatility`` = LMMSOFRFLAT from the file, ``extraSettings.includePartials`` = true
    (required for ``dataPartialDurationList`` / YBPRICE ``PDUR*`` keywords),
    ``floaterSettings`` = ``{}``, and ``globalSettings.pricingDate`` only (no ``retrievePPMProjection``).

    That shape matches ``YBPRICE(..., \"RFRSwap\", ..., \"LMMSOFRFLAT\", \"\", \"CPY\", 0, \"\")``
    plus partials when the last keyword requests PDURs.
    """
    if (sec.get("sub_type") or "").strip().lower() != "structured agency cmbs":
        return False
    if str(sec.get("yb_user_curve_ref") or "").strip():
        return False
    if not str(curve_obj.get("curveType") or "").strip():
        return False
    # Excel RFRSwap path: no embedded user curve or muni pillars on the curve object.
    if "userDefined" in curve_obj:
        return False
    return True


def run_py(session, token: str, sec: Dict[str, Any], *, level_override: Optional[float] = None) -> Dict[str, Any]:
    pricing_date = sec.get("curve_date") or default_pricing_date()
    if security_uses_municipal_curve(sec) and muni_rest_embed_curve_points_enabled():
        sec["muni_curve_points"] = resolve_muni_curve_points(session, token, sec)
    curve_obj = curve_dict_for(sec)

    prepay_type, prepay_rate = prepay_type_and_rate_for_api(sec)
    volatility = resolve_volatility(sec)
    yb_muni = muni_sync_py_ybprice_rfrswap_shape(sec, curve_obj)
    yb_sacmbs = structured_agency_cmbs_ybprice_sync_shape(sec, curve_obj)

    if yb_muni:
        # Match Excel YBPRICE(..., "RFRSwap", ..., "MatrixWSkew", ...) — minimal sync /bond/py body.
        input_row = {
            "identifier": yieldbook_security_identifier(sec),
            "floaterSettings": {},
            "extraSettings": {
                "prepayDuration": True,
                "includePartials": True,
            },
            "level": str(bond_py_level_for_request(sec, level_override)),
            "curve": curve_obj,
            "settlementDate": pricing_date,
            "volatility": volatility,
        }
        payload = {
            "globalSettings": {"pricingDate": pricing_date},
            "input": [input_row],
        }
    elif yb_sacmbs:
        # Excel YBPRICE add-in for structured agency CMBS (CUSIP.cmo, RFRSwap, LMMSOFRFLAT, CPY 0, …).
        input_row = {
            "identifier": yieldbook_security_identifier(sec),
            "floaterSettings": {},
            "extraSettings": {"includePartials": True},
            "level": str(bond_py_level_for_request(sec, level_override)),
            "curve": {"curveType": "SWAP_RFR"},
            "prepaySettings": {"type": prepay_type, "rate": prepay_rate},
            "settlementDate": pricing_date,
            "volatility": volatility,
        }
        payload = {
            "globalSettings": {"pricingDate": pricing_date},
            "input": [input_row],
        }
    else:
        input_row = {
            "identifier": yieldbook_security_identifier(sec),
            "idType": "securityIDEntry",
            "level": str(bond_py_level_for_request(sec, level_override)),
            "settlementDate": pricing_date,
            "userTag": sec["cusip"],
            "curve": curve_obj,
            "prepaySettings": {"type": prepay_type, "rate": prepay_rate},
            "volatility": volatility,
            "extraSettings": {
                "optionModel": "OASEDUR",
                # Reverse-engineered from YBPRICE partial duration behavior.
                "includePartials": True,
            },
        }
        props = py_calc_props(sec)
        if props:
            input_row["props"] = props

        payload = {
            "globalSettings": {
                "pricingDate": pricing_date,
                "retrievePPMProjection": True,
            },
            "input": [input_row],
        }

    if security_uses_municipal_curve(sec):
        points = sec.get("muni_curve_points", [])
        if sec.get("yb_user_curve_ref"):
            print(
                f"[INFO] {sec['cusip']}: Muni pricing curve=userDefined $ref "
                f"{sec['yb_user_curve_ref']!r} ({len(points)} pillar(s) uploaded to job-store)."
            )
        elif yb_muni:
            print(
                f"[INFO] {sec['cusip']}: Muni sync /bond/py = YBPRICE RFRSwap-style "
                f"(curve={curve_obj!r}, vol={volatility!r}; no prepaySettings in body)."
            )
        elif not muni_rest_embed_curve_points_enabled():
            print(
                f"[INFO] {sec['cusip']}: Muni REST curve = {curve_obj} "
                f"(no embedded muni pillars; Excel YBCURVE/YBSW may differ)."
            )
        elif points:
            print(
                f"[INFO] {sec['cusip']}: found {len(points)} muni tenor/par-rate points "
                f"(YBCURVE Municipal OnTheRun ParRate style). Pricing with {curve_obj}."
            )
        else:
            print(
                f"[INFO] {sec['cusip']}: Muni - pricing with {curve_obj} "
                f"(YB_MUNI_CURVE_TYPE overrides municipal curve enum when needed)."
            )
        if (os.environ.get("YB_PRINT_MUNI_CURVE_POINTS") or "").strip().lower() in {"1", "true", "yes", "y"}:
            print(f"[INFO] {sec['cusip']} muni_curve_points JSON:\n{muni_curve_points_debug_json(sec)}")
    elif yb_sacmbs:
        print(
            f"[INFO] {sec['cusip']}: Structured Agency CMBS /bond/py = Excel YBPRICE-style "
            f"(SWAP_RFR-only curve, prepay={prepay_type}/{prepay_rate}, vol={volatility!r}, "
            f"includePartials; id={yieldbook_security_identifier(sec)!r}).",
            flush=True,
        )

    url = api_url("bond/py", mode="sync")
    r = session.post(url, json=payload, headers=api_headers(token), timeout=60)
    if not r.ok:
        print(f"[ERROR] /bond/py failed for {sec['cusip']} HTTP {r.status_code}")
        print(f"[ERROR] pricingDate={pricing_date}, curve={curve_obj}, prepayType={prepay_type}, prepayRate={prepay_rate}, volatility={volatility}")
        print(f"[ERROR] response: {r.text[:2000]}")

        # Retry once with safest generic settings.
        retry_payload = payload.copy()
        retry_input = dict(payload["input"][0])
        retry_input["curve"] = {"curveType": "SWAP_RFR"}
        retry_input["volatility"] = {"type": "Default"}
        for k in ("prepaySettings", "idType", "userTag", "props"):
            retry_input.pop(k, None)
        if yb_muni:
            retry_input.setdefault("floaterSettings", {})
            retry_input["extraSettings"] = {"prepayDuration": True, "includePartials": True}
        else:
            retry_input["prepaySettings"] = {"type": "Model", "rate": 100}
            retry_input["curve"] = {"curveType": "SWAP_RFR", "currency": "USD"}
        retry_payload["input"] = [retry_input]
        if yb_muni:
            retry_payload["globalSettings"] = {"pricingDate": default_pricing_date()}
        else:
            retry_payload["globalSettings"] = {"pricingDate": default_pricing_date(), "retrievePPMProjection": True}

        rr = session.post(url, json=retry_payload, headers=api_headers(token), timeout=60)
        if not rr.ok:
            print(f"[ERROR] Retry /bond/py failed for {sec['cusip']} HTTP {rr.status_code}")
            print(f"[ERROR] retry response: {rr.text[:2000]}")
            rr.raise_for_status()
        py_obj = rr.json()["results"][0].get("py", {})
    else:
        py_obj = r.json()["results"][0].get("py", {})

    partial_durations = extract_partial_durations(py_obj)
    if not partial_durations:
        partial_durations = fetch_partial_durations_by_keywords(session, token, sec)
    if not partial_durations:
        print(f"[WARN] {sec['cusip']}: API did not return partial durations (including YBPRICE PDUR keywords).")

    # Normalize sync /bond/py response to expected downstream shape.
    max_servicer = (py_obj.get("maxServicer") or {})
    if isinstance(max_servicer, str):
        max_servicer = {"name": max_servicer, "percent": None}
    # YBPRICE-style flat ``py`` uses duration / dv01 / yield / convexity (not effective* names).
    eff_dur = py_obj.get("effectiveDuration")
    if eff_dur is None:
        eff_dur = py_obj.get("duration") or py_obj.get("durationToWorstCase")
    eff_conv = py_obj.get("effectiveConvexity")
    if eff_conv is None:
        eff_conv = py_obj.get("convexity")
    eff_dv01 = py_obj.get("effectiveDV01")
    if eff_dv01 is None:
        eff_dv01 = py_obj.get("dv01") or py_obj.get("dv01ToNextCall")
    flat_yield = py_obj.get("yield") or py_obj.get("yieldToWorst") or py_obj.get("semiAnnualizedYield")
    fwd_yield = (
        (py_obj.get("ForwardYield") or {}).get("Yield")
        if isinstance(py_obj.get("ForwardYield"), dict)
        else get_first_key(py_obj.get("forwardMeasures") or {}, "yield", "Yield")
    )
    if fwd_yield is None:
        fwd_yield = flat_yield
    return {
        "forwardYield": fwd_yield,
        "bondYield": flat_yield or py_obj.get("effectiveYield") or py_obj.get("streetYield"),
        "effectiveDuration": eff_dur,
        "effectiveConvexity": eff_conv,
        "effectiveDV01": eff_dv01,
        "dollarDuration": py_obj.get("dollarDuration"),
        "partialDurations": partial_durations,
        "averageLife": get_first_key(py_obj, "effectiveWAL", "EffectiveWAL", "wal", "averageLife"),
        "LongTermCPR": extract_long_term_cpr(py_obj),
        "PPHistCPRLife": extract_pphist_cpr_life(py_obj),
        "oas": py_obj.get("oas"),
        "zSpread": py_obj.get("zSpread"),
        "factor": get_first_key(py_obj, "factor", "Factor"),
        "GrossWAC": get_first_key(py_obj, "grossWAC", "GrossWAC", "gwac", "GWAC"),
        "LoanAge": get_first_key(py_obj, "loanAge", "LoanAge", "wala", "WALA"),
        "WeightedAvgLoanSize": get_first_key(py_obj, "weightedAvgLoanSize", "WeightedAvgLoanSize", "wals", "WALS"),
        "maxServicer": max_servicer,
    }

# ---------------------------------------------------------
# Scenario Calc
# ---------------------------------------------------------
def run_scenarios(session, token: str, sec: Dict[str, Any]) -> List[Dict[str, Any]]:
    pricing_date = sec.get("curve_date") or default_pricing_date()
    if (
        security_uses_municipal_curve(sec)
        and muni_rest_embed_curve_points_enabled()
        and not (sec.get("muni_curve_points") or [])
    ):
        sec["muni_curve_points"] = resolve_muni_curve_points(session, token, sec)
    scen_curve = curve_dict_for(sec)
    scen_volatility = resolve_volatility(sec)

    scenarios = []

    prepay_type, prepay_rate = prepay_type_and_rate_for_api(sec)
    for i, s in enumerate(SHOCKS_BPS, start=1):
        scenarios.append({
            "scenarioID": f"scen{i}",
            "timing": scenario_timing(),
            "definition": {
                "userScenario": {
                    "shiftType": SCENARIO_SHIFT_TYPE,
                    "interpolationType": SCENARIO_INTERPOLATION_TYPE,
                    "curveShifts": [{"year": 0.25, "value": s}],
                }
            }
        })
    payload = {
        "globalSettings": {
            "pricingDate": pricing_date,
            "horizonDays": SCENARIO_HORIZON_DAYS,
            "horizonMonths": SCENARIO_HORIZON_MONTHS,
        },
        "scenarios": scenarios,
        "calcInputs": [{
            "userTag": sec["cusip"],
            "identifier": yieldbook_security_identifier(sec),
            "idType": "securityIDEntry",
            **({"props": py_calc_props(sec)} if py_calc_props(sec) else {}),
            "curve": scen_curve,
            "volatility": scen_volatility,
            "settlementInfo": {
                "level": bond_py_level_value(sec),
                "settlementType": "CUSTOM",
                "settlementDate": pricing_date,
                "prepay": {
                    "type": prepay_type,
                    "rate": prepay_rate,
                },
            },
            # Match settlement prepay: horizon legs previously sent rate-only and could diverge from Excel.
            "horizonInfo": [
                {
                    "scenarioID": f"scen{i}",
                    "level": 0,
                    "prepay": {"type": prepay_type, "rate": prepay_rate},
                }
                for i, _ in enumerate(SHOCKS_BPS, start=1)
            ],
            "assumeCall": False,
            "horizonPYMethod": scenario_horizon_py_method(),
        }],
    }

    #url = f"{API_BASE}/req/bond/scenario-calc"
    url = api_url("bond/scenario-calc", mode="req")
    r = session.post(url, json=payload, headers=api_headers(token), timeout=120)
    if not r.ok:
        print(f"[ERROR] /bond/scenario-calc failed for {sec['cusip']} HTTP {r.status_code}")
        print(f"[ERROR] pricingDate={pricing_date}, curve={scen_curve}, prepayType={prepay_type}, prepayRate={prepay_rate}")
        print(f"[ERROR] response: {r.text[:2000]}")
        return []
    data = r.json()
    request_id = data.get("requestId")
    if not request_id:
        print(f"[WARN] scenario-calc: no requestId for {sec['cusip']}: {data}")
        return []

    results_url = api_url(f"/results/{request_id}", mode=None)
    last_bad_status: Optional[int] = None
    for attempt in range(24):
        rr = session.get(results_url, headers=api_headers(token), timeout=30)
        if rr.status_code == 404:
            time.sleep(2)
            continue
        if not rr.ok:
            if rr.status_code != last_bad_status:
                print(
                    f"[WARN] scenario results GET {request_id} HTTP {rr.status_code} "
                    f"for {sec['cusip']} (will retry): {(rr.text or '')[:300]}"
                )
                last_bad_status = rr.status_code
            time.sleep(2)
            continue
        try:
            jr = rr.json()
        except Exception:
            time.sleep(2)
            continue
        meta = jr.get("meta") or {}
        if meta.get("status") == "DONE":
            results = jr.get("results", [])
            if not results:
                return []
            raw_h = (results[0].get("scenario") or {}).get("horizon", [])
            return normalize_scenario_horizon(raw_h)
        if meta.get("status") == "ERROR":
            print(
                f"[WARN] scenario results ERROR for {sec['cusip']} request {request_id}: "
                f"{jr.get('errors') or meta}"
            )
            return []
        time.sleep(2)
    print(f"[WARN] Timed out waiting for scenario results for {sec['cusip']} ({request_id})")
    return []


def pick_metric(h: Dict[str, Any], *keys: str) -> Any:
    for k in keys:
        if k in h and h.get(k) is not None:
            return h.get(k)
    return None


def derive_effective_dv01_at_horizon(h: Dict[str, Any]) -> Optional[float]:
    """
    Fallback approximation when API does not return EffectiveDV01AtHorizon.
    DV01 ~= Duration * Price / 10000.
    """
    dur = pick_metric(h, *SCENARIO_EFFDUR_KEYS)
    px = pick_metric(h, "fullPrice", "price", "actualFullPrice", "actualPrice")
    if dur is None or px is None:
        return None
    try:
        return float(dur) * float(px) / 10000.0
    except (TypeError, ValueError):
        return None


def merge_scenario_into_row(row: Dict[str, Any], scen: List[Dict[str, Any]]) -> Dict[str, Any]:
    for i, shock in enumerate(SHOCKS_BPS):
        h = scen[i] if i < len(scen) else {}
        shock_label = f"{shock:+d}" if shock > 0 else str(shock)
        # Yield Book scenario horizon typically exposes `duration` (risk duration after shock).
        row[f"EffDur_{shock_label}"] = pick_metric(h, *SCENARIO_EFFDUR_KEYS)
        dv01_val = pick_metric(
            h,
            "effectiveDV01AtHorizon",
            "dv01AtHorizon",
            "effectiveDV01",
            "dv01",
            "spreadDV01",
        )
        dv01_num: Optional[float] = None
        try:
            if dv01_val is not None:
                dv01_num = float(dv01_val)
        except (TypeError, ValueError):
            dv01_num = None
        if dv01_num is None or abs(dv01_num) < 1e-12:
            derived_dv01 = derive_effective_dv01_at_horizon(h)
            if derived_dv01 is not None:
                dv01_val = derived_dv01
        row[f"DV01_{shock_label}"] = dv01_val
        # User rule: use API dollarReturn only for final shock results.
        row[f"DollarReturn_{shock_label}"] = h.get("dollarReturn")
    return row


def extract_partial_durations(py_obj: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract partial/key-rate durations from common response shapes.
    Returns keys normalized to: 1Y, 2Y, 3Y, 5Y, 10Y, 20Y, 30Y.
    """
    candidates = [
        py_obj.get("partialDurations"),
        py_obj.get("partialDuration"),
        py_obj.get("keyRateDurations"),
        py_obj.get("keyrateDurations"),
        py_obj.get("keyRateDuration"),
    ]

    valid = {"1Y", "2Y", "3Y", "5Y", "10Y", "20Y", "30Y"}

    for c in candidates:
        if not c:
            continue
        if isinstance(c, dict):
            out: Dict[str, Any] = {}
            for k, v in c.items():
                k2 = str(k).strip().upper().replace("YR", "Y")
                if k2 in {"1", "2", "3", "5", "10", "20", "30"}:
                    k2 = f"{k2}Y"
                if k2 in valid:
                    out[k2] = v
            if out:
                return out
        if isinstance(c, list):
            out = {}
            for item in c:
                if not isinstance(item, dict):
                    continue
                node = pick_metric(item, "node", "year", "tenor", "key")
                val = pick_metric(item, "value", "duration", "partialDuration")
                if node is None or val is None:
                    continue
                try:
                    key = f"{int(float(node))}Y"
                except Exception:
                    continue
                if key in valid:
                    out[key] = val
            if out:
                return out
    # Common sync/bond/py shape when extraSettings.includePartials = true
    c = py_obj.get("dataPartialDurationList")
    if isinstance(c, list):
        out = {}
        for item in c:
            if not isinstance(item, dict):
                continue
            node = pick_metric(item, "partialDurationYear", "year", "node")
            val = pick_metric(item, "partialDuration", "duration", "value")
            if node is None or val is None:
                continue
            try:
                key = f"{int(float(node))}Y"
            except Exception:
                continue
            if key in {"1Y", "2Y", "3Y", "5Y", "10Y", "20Y", "30Y"}:
                out[key] = val
        if out:
            return out
    return {}


def extract_partial_durations_from_keyword_py(py_obj: Dict[str, Any]) -> Dict[str, Any]:
    """
    Extract partial durations from keyword-based YBPRICE style field names.
    Examples: PartialDuration1year, partialDuration10Year, PDUR1..PDUR7.
    """
    out: Dict[str, Any] = {}
    if not isinstance(py_obj, dict):
        return out

    year_map = {
        "1": "1Y",
        "2": "2Y",
        "3": "3Y",
        "5": "5Y",
        "10": "10Y",
        "20": "20Y",
        "30": "30Y",
    }
    pdur_map = {
        "PDUR1": "1Y",
        "PDUR2": "2Y",
        "PDUR3": "3Y",
        "PDUR4": "5Y",
        "PDUR5": "10Y",
        "PDUR6": "20Y",
        "PDUR7": "30Y",
    }

    for k, v in py_obj.items():
        ku = str(k).strip().upper()
        if ku in pdur_map and v is not None:
            out[pdur_map[ku]] = v
            continue
        if "PARTIALDURATION" in ku:
            # Handle forms like PartialDuration1year / PartialDuration10Year
            digits = "".join(ch for ch in ku if ch.isdigit())
            if digits in year_map and v is not None:
                out[year_map[digits]] = v
    return out


def fetch_partial_durations_by_keywords(session, token: str, sec: Dict[str, Any]) -> Dict[str, Any]:
    """
    Try YBPRICE-style keyword request for partial durations via req/bond/py.
    """
    pricing_date = sec.get("curve_date") or default_pricing_date()
    if (
        security_uses_municipal_curve(sec)
        and muni_rest_embed_curve_points_enabled()
        and not (sec.get("muni_curve_points") or [])
    ):
        sec["muni_curve_points"] = resolve_muni_curve_points(session, token, sec)
    curve_obj = curve_dict_for(sec)
    prepay_type, prepay_rate = prepay_type_and_rate_for_api(sec)
    scen_vol = resolve_volatility(sec)

    keywords = [
        "PartialDuration1year",
        "PartialDuration2year",
        "PartialDuration3year",
        "PartialDuration5year",
        "PartialDuration10year",
        "PartialDuration20year",
        "PartialDuration30year",
        "PDUR1",
        "PDUR2",
        "PDUR3",
        "PDUR4",
        "PDUR5",
        "PDUR6",
        "PDUR7",
    ]

    py_in: Dict[str, Any] = {
            "identifier": yieldbook_security_identifier(sec),
            "idType": "securityIDEntry",
            "level": str(bond_py_level_value(sec)),
            "userTag": sec["cusip"],
            "curve": curve_obj,
            "prepaySettings": {"type": prepay_type, "rate": prepay_rate},
            "volatility": scen_vol,
            "extraSettings": {"optionModel": "OASEDUR"},
    }
    props = py_calc_props(sec)
    if props:
        py_in["props"] = props

    body = {
        "keywords": keywords,
        "globalSettings": {"pricingDate": pricing_date},
        "pyCalcInputs": [py_in],
    }

    r = session.post(api_url("bond/py", mode="req"), json=body, headers=api_headers(token), timeout=60)
    if not r.ok:
        return {}
    request_id = r.json().get("requestId")
    if not request_id:
        return {}

    results_url = api_url(f"/results/{request_id}", mode=None)
    for _ in range(20):
        rr = session.get(results_url, headers=api_headers(token), timeout=30)
        if rr.status_code == 404:
            time.sleep(1)
            continue
        if not rr.ok:
            return {}
        jr = rr.json()
        if jr.get("meta", {}).get("status") == "DONE":
            py_kw = (jr.get("results") or [{}])[0].get("py", {})
            return extract_partial_durations_from_keyword_py(py_kw)
        time.sleep(1)
    return {}


def run_indic(session, token: str, sec: Dict[str, Any]) -> Dict[str, Any]:
    """
    YBINDIC-equivalent pull via REST /sync/bond/indic.
    Requested fields:
    Factor, PPHistCPRLife, GrossWAC, LoanAge, WeightedAvgLoanSize,
    MaxServiceName/MaxServicerName, MaxServicerPercent.
    """
    indic_date = sec.get("curve_date") or default_pricing_date()
    keywords = [
        "Factor",
        "PPHistCPRLife",
        "GrossWAC",
        "LoanAge",
        "WeightedAvgLoanSize",
        "MaxServiceName",      # keep user-requested spelling
        "MaxServicerName",     # also try API-servicer spelling
        "MaxServicerPercent",
    ]
    def _request_indic(use_keywords: bool) -> Dict[str, Any]:
        payload = {
            "identifierInfos": [{
                "identifier": yieldbook_security_identifier(sec),
                "idType": "securityIDEntry",
            }],
            "globalSettings": {
                "indicDate": indic_date,
            },
        }
        if use_keywords:
            payload["keywords"] = keywords
        resp = session.post(
            api_url("bond/indic", mode="sync"),
            json=payload,
            headers=api_headers(token),
            timeout=60,
        )
        if not resp.ok:
            if use_keywords:
                print(f"[WARN] /bond/indic failed for {sec['cusip']} HTTP {resp.status_code}: {resp.text[:500]}")
            return {}
        results = resp.json().get("results") or []
        return results[0].get("indic") or {} if results else {}

    indic = _request_indic(use_keywords=True)
    if not indic:
        return {}

    def _extract_values(indic_obj: Dict[str, Any]) -> Dict[str, Any]:
        collateral_list = indic_obj.get("dataCollateralList")
        collateral = collateral_list[0] if isinstance(collateral_list, list) and collateral_list else {}
        if not isinstance(collateral, dict):
            collateral = {}

        gross_wac = get_first_key(indic_obj, "grossWAC", "GrossWAC", "gwac", "GWAC")
        loan_age = get_first_key(indic_obj, "loanAge", "LoanAge", "wala", "WALA")
        wals = get_first_key(indic_obj, "weightedAvgLoanSize", "WeightedAvgLoanSize", "wals", "WALS")
        if gross_wac is None:
            gross_wac = get_first_key(collateral, "grossWAC", "GrossWAC", "gwac", "GWAC")
        if loan_age is None:
            loan_age = get_first_key(collateral, "loanAge", "LoanAge", "wala", "WALA")
        if wals is None:
            wals = get_first_key(collateral, "weightedAvgLoanSize", "WeightedAvgLoanSize", "wals", "WALS", "loanSize", "LoanSize")

        max_name = get_first_key(indic_obj, "maxServicerName", "MaxServicerName", "maxServiceName", "MaxServiceName")
        max_pct = get_first_key(indic_obj, "maxServicerPercent", "MaxServicerPercent")
        if max_name is None or max_pct is None:
            serv_list = collateral.get("maxServicerList")
            if isinstance(serv_list, list) and serv_list and isinstance(serv_list[0], dict):
                top_serv = serv_list[0]
                if max_name is None:
                    max_name = top_serv.get("name")
                if max_pct is None:
                    max_pct = top_serv.get("percent")

        return {
            "factor": get_first_key(indic_obj, "factor", "Factor"),
            "PPHistCPRLife": extract_pphist_cpr_life(indic_obj),
            "GrossWAC": gross_wac,
            "LoanAge": loan_age,
            "WeightedAvgLoanSize": wals,
            "maxServicerName": max_name,
            "maxServicerPercent": max_pct,
        }

    extracted = _extract_values(indic)
    if (
        extracted["GrossWAC"] is None
        and extracted["LoanAge"] is None
        and extracted["WeightedAvgLoanSize"] is None
        and not indic.get("dataCollateralList")
    ):
        indic_full = _request_indic(use_keywords=False)
        if indic_full:
            extracted = _extract_values(indic_full)
    elif extracted.get("PPHistCPRLife") is None:
        # Keyword-filtered /bond/indic often omits dataPPMHistoryList; Life CPR lives in CPR history (month Life).
        indic_full = _request_indic(use_keywords=False)
        if indic_full:
            ext_full = _extract_values(indic_full)
            if ext_full.get("PPHistCPRLife") is not None:
                extracted["PPHistCPRLife"] = ext_full["PPHistCPRLife"]

    return extracted


def run_py_keyword_measures(session, token: str, sec: Dict[str, Any]) -> Dict[str, Any]:
    """
    YBPRICE keyword-style pull for:
    EffectiveWAL, LongTermCPR, OAS, ZSpread.
    """
    pricing_date = sec.get("curve_date") or default_pricing_date()
    sub_type = (sec.get("sub_type") or "").strip()
    prepay_type, prepay_rate = prepay_type_and_rate_for_api(sec)
    # Optional override for Agency MBS ARM benchmarking; default is to respect input prepay rate.
    arm_force_raw = (os.environ.get("YB_ARM_FORCE_PREPAY_RATE") or "").strip()
    if sub_type.lower() == "agency mbs arm" and arm_force_raw:
        arm_force = parse_number(arm_force_raw)
        if arm_force is not None:
            prepay_rate = float(arm_force)
    if security_uses_municipal_curve(sec) and muni_rest_embed_curve_points_enabled():
        sec["muni_curve_points"] = resolve_muni_curve_points(session, token, sec)
    curve_obj = curve_dict_for(sec)

    def _submit(volatility: Dict[str, Any]) -> Dict[str, Any]:
        body = {
            "keywords": ["EffectiveWAL", "LongTermCPR", "OAS", "ZSpread"],
            "globalSettings": {"pricingDate": pricing_date},
            "pyCalcInputs": [{
                "identifier": yieldbook_security_identifier(sec),
                "idType": "securityIDEntry",
                "level": str(bond_py_level_value(sec)),
                "settlementDate": pricing_date,
                # REST does not expose a direct securitySubType field, use props to pass subtype context.
                "props": py_calc_props(sec) or {"subType": sub_type} if sub_type else {},
                "curve": curve_obj,
                "prepaySettings": {"type": prepay_type, "rate": prepay_rate},
                "volatility": volatility,
                "extraSettings": {"optionModel": "OASEDUR"},
            }],
        }
        r = session.post(api_url("bond/py", mode="req"), json=body, headers=api_headers(token), timeout=60)
        if not r.ok:
            return {}
        request_id = r.json().get("requestId")
        if not request_id:
            return {}
        results_url = api_url(f"/results/{request_id}", mode=None)
        for _ in range(20):
            rr = session.get(results_url, headers=api_headers(token), timeout=30)
            if rr.status_code == 404:
                time.sleep(1)
                continue
            if not rr.ok:
                return {}
            jr = rr.json()
            if jr.get("meta", {}).get("status") == "DONE":
                return (jr.get("results") or [{}])[0].get("py", {}) or {}
            time.sleep(1)
        return {}

    # Try resolved volatility first, then fallback to Default.
    primary_vol = resolve_volatility(sec)
    py_kw = _submit(primary_vol)
    if py_kw.get("returnCode") == 1 or not py_kw:
        if str(primary_vol.get("type", "")).strip().upper() != "DEFAULT":
            py_kw = _submit({"type": "Default"})

    return {
        "averageLife": py_kw.get("effectiveWAL"),
        "LongTermCPR": extract_long_term_cpr(py_kw),
        "oas": py_kw.get("oas"),
        "zSpread": py_kw.get("zSpread"),
    }


def run_full_analysis_for_security(session, token: str, sec: Dict[str, Any]) -> Dict[str, Any]:
    """One security: PY + keyword + scenario + indic; merged output row for OUTPUT_COLUMNS."""
    py = run_py(session, token, sec)
    book_lvl = parse_number(sec.get("book_price"))
    py_book: Optional[Dict[str, Any]] = None
    if book_lvl is not None and not pd.isna(book_lvl):
        try:
            py_book = run_py(session, token, sec, level_override=float(book_lvl))
        except Exception as e:
            print(f"[WARN] {sec['cusip']}: Prospective_Yield book-price /bond/py skipped: {type(e).__name__}: {e}")
    py_kw = run_py_keyword_measures(session, token, sec)
    scen = run_scenarios(session, token, sec)
    indic = run_indic(session, token, sec)

    row = {
        "CUSIP": sec["cusip"],
        "Sub Type": sec.get("sub_type"),

        "Forward_Yield": resolve_forward_yield_column(py, sec.get("sub_type")),
        "Prospective_Yield": resolve_prospective_yield_from_py(py_book, sec.get("sub_type")),
        "Effective_Duration": py.get("effectiveDuration"),
        "Effective_Convexity": py.get("effectiveConvexity"),
        "Effective_DV01": py.get("effectiveDV01"),
        "Dollar_Duration": computed_dollar_duration(sec, py),

        "PD_1Y": py.get("partialDurations", {}).get("1Y"),
        "PD_2Y": py.get("partialDurations", {}).get("2Y"),
        "PD_3Y": py.get("partialDurations", {}).get("3Y"),
        "PD_5Y": py.get("partialDurations", {}).get("5Y"),
        "PD_10Y": py.get("partialDurations", {}).get("10Y"),
        "PD_20Y": py.get("partialDurations", {}).get("20Y"),
        "PD_30Y": py.get("partialDurations", {}).get("30Y"),

        "Average_Life": py_kw.get("averageLife") if py_kw.get("averageLife") is not None else py.get("averageLife"),
        "LT_CPR": py_kw.get("LongTermCPR") if py_kw.get("LongTermCPR") is not None else py.get("LongTermCPR"),
        "Life_CPR": resolve_life_cpr_value(sec.get("sub_type"), py, indic),
        "OAS": py_kw.get("oas") if py_kw.get("oas") is not None else py.get("oas"),
        "Z_Spread": py_kw.get("zSpread") if py_kw.get("zSpread") is not None else py.get("zSpread"),
        "Factor": py.get("factor") if py.get("factor") is not None else indic.get("factor"),
        "GWAC": py.get("GrossWAC") if py.get("GrossWAC") is not None else indic.get("GrossWAC"),
        "WALA": py.get("LoanAge") if py.get("LoanAge") is not None else indic.get("LoanAge"),
        "WALS": py.get("WeightedAvgLoanSize") if py.get("WeightedAvgLoanSize") is not None else indic.get("WeightedAvgLoanSize"),
        "MaxServicerName": py.get("maxServicer", {}).get("name")
        if py.get("maxServicer", {}).get("name") is not None
        else indic.get("maxServicerName"),
        "MaxServicerPercent": py.get("maxServicer", {}).get("percent")
        if py.get("maxServicer", {}).get("percent") is not None
        else indic.get("maxServicerPercent"),
    }

    row = merge_scenario_into_row(row, scen)
    return row


def run_fast_scenario_refresh(session, token: str, securities: List[Dict[str, Any]]) -> None:
    """
    Refresh only scenario shock columns in existing OUTPUT_CSV.
    """
    scen_cols = scenario_columns()
    if os.path.isfile(OUTPUT_CSV):
        df = pd.read_csv(OUTPUT_CSV)
    else:
        df = pd.DataFrame(columns=["CUSIP"] + scen_cols)

    if "CUSIP" not in df.columns:
        raise ValueError(f"{OUTPUT_CSV} must contain a CUSIP column for fast refresh mode.")
    if "Sub Type" not in df.columns:
        df["Sub Type"] = pd.NA

    for c in scen_cols:
        if c not in df.columns:
            df[c] = pd.NA

    # Normalize join key
    df["CUSIP"] = df["CUSIP"].astype(str).str.strip()
    idx_by_cusip = {c: i for i, c in enumerate(df["CUSIP"].tolist())}

    for sec in securities:
        cusip = str(sec.get("cusip", "")).strip()
        sub_type = sec.get("sub_type")
        if not cusip:
            continue
        try:
            scen = run_scenarios(session, token, sec)
        except Exception as e:
            print(f"[ERROR] Scenario refresh failed for {cusip}: {type(e).__name__}: {e}")
            continue

        upd = merge_scenario_into_row({"CUSIP": cusip}, scen)
        if cusip in idx_by_cusip:
            row_idx = idx_by_cusip[cusip]
            df.at[row_idx, "Sub Type"] = sub_type
            for c in scen_cols:
                df.at[row_idx, c] = upd.get(c)
        else:
            new_row = {col: pd.NA for col in df.columns}
            new_row["CUSIP"] = cusip
            new_row["Sub Type"] = sub_type
            for c in scen_cols:
                new_row[c] = upd.get(c)
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            idx_by_cusip[cusip] = len(df) - 1

    df.to_csv(OUTPUT_CSV, index=False, float_format="%.6f")
    print(f"Fast scenario refresh complete: {OUTPUT_CSV}")

# ---------------------------------------------------------
# Main
# ---------------------------------------------------------
def main():
    session = make_http_session()
    token = get_access_token(session)

    securities = load_securities()
    fast_mode = str(os.getenv("FAST_SCENARIO_ONLY", "")).strip().lower() in {"1", "true", "yes", "y"}
    if fast_mode:
        run_fast_scenario_refresh(session, token, securities)
        return

    workers = parallel_worker_count()
    rows: List[Dict[str, Any]] = []

    if workers > 1:
        print(f"[INFO] YB_WORKERS={workers} (parallel securities).")

        def _work(sec: Dict[str, Any]) -> Optional[Dict[str, Any]]:
            s = make_http_session()
            try:
                return run_full_analysis_for_security(s, token, sec)
            except Exception as e:
                print(f"[ERROR] Skipping {sec.get('cusip')} due to API error: {type(e).__name__}: {e}")
                return None

        with ThreadPoolExecutor(max_workers=workers) as ex:
            out = list(ex.map(_work, securities))
        rows = [r for r in out if r is not None]
    else:
        for sec in securities:
            try:
                rows.append(run_full_analysis_for_security(session, token, sec))
            except Exception as e:
                print(f"[ERROR] Skipping {sec.get('cusip')} due to API error: {type(e).__name__}: {e}")
                continue

    df = pd.DataFrame(rows)

    # Enforce exact Excel template layout
    df = df.reindex(columns=OUTPUT_COLUMNS)

    df.to_csv(OUTPUT_CSV, index=False, float_format="%.6f")
    print(f"Saved results to {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
